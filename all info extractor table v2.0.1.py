import matplotlib.pyplot as plt
import openpyxl
from openpyxl.styles import Font, Color, Alignment, PatternFill
from openpyxl.styles import colors
from openpyxl.styles.borders import Border, Side
from openpyxl.worksheet import worksheet
import pandas as pd
import numpy as np
from pandas import ExcelWriter
from pandas import ExcelFile
from PIL import Image
import time
import sys

"""
A PROGRAM TO EXTRACT DATA AND LABEL DATA BASED ON ONE CHARACTERISTIC.
PROVIDE IV CURVE OF THE CHAMPION CELL.
"""

#COLOR FUNCTIONS --------------------------------------------------------------

def color_range(row_num, chosen_cat):
    """
    Find the rows in resulting excel table to color all info about best cell
    """
    color_range_best = []
    if chosen_cat == 0: #eff
        color_range_best = list(range(row_num, row_num + 4))
    elif chosen_cat == 1: #ff
        color_range_best = [row_num - 1, row_num, row_num + 1, row_num + 2]
    elif chosen_cat == 2: #voc
        color_range_best = [row_num - 2, row_num - 1, row_num, row_num + 1]
    
    return color_range_best

def row_num_diff(row_num, chosen_cat):
    """
    Find the row of voltage based on the row_num
    """
    if chosen_cat == 0: #eff
        v_row = row_num - 2
    elif chosen_cat == 1:
        v_row = row_num - 3
    elif chosen_cat == 2:
        v_row = row_num -4
    
    return v_row

def color_eff(val):
    """
    Color for Efficiency.
    """    
    val = float(val)
    color = "white"
    if val < 1:
        color = "#ff6961" #Red
    elif val < 3:
        color = "#fdfd96" #Yellow
    elif val < 6:
        color = "#a0e7a0" #Green
    elif val > 6:
        color = "#cde7f0" #Blue
    else: 
        color = "white"
    return 'background-color: %s' % color  

def color_ff(val):
    """
    Color for Fill Factor.
    """    
    val = float(val)
    color = "white"
    if val < 0.25:
        color = "#ff6961" #Red
    elif val < 0.35:
        color = "#fdfd96" #Yellow
    elif val < 0.4:
        color = "#a0e7a0" #Green
    elif val > 0.4:
        color = "#cde7f0" #Blue
    else: 
        color = "white"
    return 'background-color: %s' % color 

def color_voc(val):
    """
    Color for Voc.
    """    
    val = float(val)
    color = "white"
    if val < 0.8:
        color = "#ff6961" #Red
    elif val < 0.9:
        color = "#fdfd96" #Yellow
    elif val < 1:
        color = "#a0e7a0" #Green
    elif val > 1:
        color = "#cde7f0" #Blue
    else: 
        color = "white"
    return 'background-color: %s' % color 

def color_jsc(value):
    if pd.isnull(value):
        color = "white"
    else:
        color = "#ffebef"
    return 'background-color: %s' % color

all_color_functions = [color_eff, color_ff, color_voc]

def leg_colors(value):
    """
    Color for legend.
    """
    if value == "Red":
        color = "#ff6961"
    elif value == "Yellow":
        color = "#fdfd96"
    elif value == "Green":
        color = "#a0e7a0"
    elif value == "Blue":
        color = "#cde7f0"
    elif value == "Purple":
        color = "#e3c7e9" 
    elif value == "White":
        color = "white"
        
    return 'background-color: %s' % color

#HELPER FUNCTIONS --------------------------------------------------------------
def find_index(value, dataf, cat):
    """
    Find the index of the value in the dataframe
    df coor example [(1, "Eff"), "B"]
    """
    cat = cat.lower()
    if cat == "eff":
        cat_id_list = eff_subdf_id
    elif cat == "ff":
        cat_id_list = ff_subdf_id
    elif cat == "voc":
        cat_id_list = voc_subdf_id
    elif cat == "jsc":
        cat_id_list = jsc_subdf_id
        
    for index in cat_id_list: 
        for col in cols:
            if value == dataf.loc[index, col]:
                return [index, col]

def xl_coordinate(df_coor, df):
    """
    Convert df_coor into xl_coor. Return a string "rowcol".
    """
    eff_xl_rows = list(range(2, 23, 4))
    ff_xl_rows = list(range(3,24,4))
    voc_xl_rows = list(range(4,25,4))
    jsc_xl_rows = list(range(5,26,4))
    
    all_xl_rows = [eff_xl_rows, ff_xl_rows, voc_xl_rows, jsc_xl_rows]
    
    category = df_coor[0][1]
    
    index = row_cat.index(category)
    
    cat_xl_rows = all_xl_rows[index]
    
    xl_row = str(cat_xl_rows[df_coor[0][0] - 1])
    xl_col = chr(ord(df_coor[1]) + 2)
    xl_coor = xl_col+xl_row
    return xl_coor

def get_export_filename(xl_source):
    """
    Get the name of cell and the date measured. Assume the name from the first
    cell measured and the date is from the last cell measured in the merged
    exel file.
    """
    wb = openpyxl.load_workbook(xl_source)
    sheet = wb.worksheets[0]
    
    #Get The Time Measured
    last_row_num = sheet.max_row
    
    #Most recent date indicates the data retrieved
    most_recent_date_row = last_row_num - 11
    
    #Get the date 
    date_measured = sheet["A" + str(most_recent_date_row)].value
    
    #Convert the date into string and get only the date
    date_str = str(date_measured).split()[0]     
    
    #Get the name of batch
    batch_name = sheet['B2'].value[3:]
    
    out_name = batch_name + " measured " + date_str + " table " + \
        cat_list[chosen_cat] + ".xlsx"
    
    return out_name

def create_legend_df():
    """
    Create the legend for the data
    """
    label_colors = ['Red', 'Yellow', 'Green', 'Blue', 'Purple', 'White']
    eff_values = ['< 1', '< 3', '< 6', '> 6', 'best', 'empty']
    ff_values = ['< 0.25', '< 0.35', '< 0.4', '> 0.4', 'best', 'empty']
    voc_values = ['< 0.8', '< 0.9', '< 1', '> 1', 'best', 'empty']
    jsc_values = ['', '', '', '', '']
    
    all_cat_names = ['Eff(%)', 'FF', 'Voc (V)', 'Jsc (mA/mm^2)']
    all_values = [eff_values, ff_values, voc_values, jsc_values]
    
    legend = {'Color': label_colors, all_cat_names[chosen_cat]: \
              all_values[chosen_cat]}
    
    df_legend = pd.DataFrame(legend)
    df_legend_style = df_legend.style.applymap(leg_colors, subset = ['Color'])
    
    return [df_legend,df_legend_style]

#EXECUTION --------------------------------------------------------------------
#PROMPT USER INPUT ------------------------------------------------------------
ques = "Have you converted the txt file into an excel file and saved it? [Y/N] "
ans = input(ques)

if ans != "Y" and ans != "y":
    print("Convert it and hit play button again.")
    sys.exit("Convert it and hit play button again.")

xl_sourcefile = input("What is the name of the xl source file? ") + ".xlsx"

#Prompt the user to choose a category to color the table. 
#chosen_cat is a number 

print("Choose a characterisitic to color the table.")
print("[0]eff [1]ff [2]voc")

chosen_cat = input("Please input a number 0, 1, 2. ")
while not chosen_cat.isdigit():
    print("Please input a number only.")
    chosen_cat = input("Please input a number 0, 1, 2. ")

chosen_cat = int(chosen_cat)

while chosen_cat not in [0, 1, 2]:
    print("Please only input 0, 1 or 2.")
    chosen_cat = int(input("Please only input 0, 1 or 2. "))

start_time = time.time()

#EXTRACT DATA AND PUT IN DATAFRAME --------------------------------------------
#Open the excel file
wb = openpyxl.load_workbook(xl_sourcefile)
sheet = wb.worksheets[0]    
  
data = pd.read_excel(xl_sourcefile)

#Basic information
cat_list = ["eff", "ff", "voc", "jsc"]
row_id = list(range(1,7))
row_cat = ['Eff', 'Ff', 'Voc', 'Jsc']
units = ['%', '', 'V', 'mA/mm^2']
cols = list('ABCDEF')

name_list = []
eff_index_list = [] #indexing in the dataframe 
eff_list = []

ff_index_list = [] #indexing in the dataframe 
ff_list = []

voc_index_list = [] #indexing in the dataframe 
voc_list = []

isc_index_list = [] #indexing in the dataframe 
isc_list = []

area_index_list = []
area_list = [] 
jsc_list = []

#Create a list of every item of forward and reverse. Compare the two items
#together, and update the list for the item being chosen.

for index, row in data.iterrows():
    l_row = list(row)
    if "Comment:" in l_row:
        if  "d" not in l_row[1]:
            #"Reverse" not in l_row[1] and
            #l_row[1] is the name of the cell
            name_list.append(l_row[1])
            
            eff_index = index + 3
            eff_index_list.append(eff_index)
            
            ff_index = index + 4
            ff_index_list.append(ff_index)
            
            voc_index = index + 5
            voc_index_list.append(voc_index)
            
            isc_index = index + 6
            isc_index_list.append(isc_index)    
            
            area_index = index + 9
            area_index_list.append(area_index)


#Process through the information based on the characterisitc chosen by the user
all_index_list = [eff_index_list, ff_index_list, voc_index_list, \
                  isc_index_list]
all_index_list_id = list(range(len(all_index_list)))

#Create a list that gets information 
#Keep an index list (index = index in the list of row numbers in the raw df) 
#that shows that the data is reverse data

cat_index_off = [] #index is row number in raw dataframe
chosen_index_list = []
#chose_index_list records the list index of data that are greater
#even number is forward, odd number is reverse

#iterate through the list of data of a characteristic chosen by user and 
#determine if forward and reverse data is better.
#Synchronize the data by updating the other index lists accordingly

#Indexing in excel is different 
xl_eff_list = [str(i+2) for i in eff_index_list] 
xl_ff_list = [str(i+2) for i in ff_index_list] 
xl_voc_list = [str(i+2) for i in voc_index_list]             
xl_isc_list = [str(i+2) for i in isc_index_list]
xl_area_list = [str(i+2) for i in area_index_list]

all_xl_list = [xl_eff_list, xl_ff_list, xl_voc_list, xl_isc_list]
all_cat_list = [eff_list, ff_list, voc_list, isc_list]

for i in range(len(all_xl_list) - 1):
    for row_num in all_xl_list[i]:
        all_cat_list[i].append(sheet['B' + row_num].value) 
        
for row_num in xl_isc_list:
    isc_list.append(sheet['B' + row_num].value*1000)
    #change unit into mA
    
for row_num in xl_area_list:
    area_list.append(sheet['B' + row_num].value)

for i in range(len(isc_list)):
    jsc = round(isc_list[i]/area_list[i],3)
    jsc_list.append(jsc)

all_idx = pd.MultiIndex.from_product([row_id, row_cat])
df = pd.DataFrame("", all_idx, cols)

all_data = [eff_list, ff_list, voc_list, jsc_list, name_list]
all_data_index = list(range(len(all_data)))

#Process the data
chosen_cat_list = all_data[chosen_cat]
chosen_index_list = []
#if element in chosen index list is odd - reverse data, even - forward data

#cat_list_official
cat_list_off = []

for i in range(0, len(chosen_cat_list), 2):    
    if chosen_cat_list[i] >= chosen_cat_list[i+1]:
    #if forward data is better
        chosen_index_list.append(i)
        cat_list_off.append(chosen_cat_list[i])
    elif chosen_cat_list[i] < chosen_cat_list[i+1]:
        chosen_index_list.append(i+1)
        cat_list_off.append(chosen_cat_list[i+1])

#all data after being processed
all_data_off = [[],[],[],[],[]]
#order: eff, ff, voc, jsc, name 

all_data_off[chosen_cat] = cat_list_off
all_data_index.remove(chosen_cat)

for index in all_data_index:
    for idx in chosen_index_list:
        all_data_off[index].append(all_data[index][idx])

name_list_off = all_data_off[-1]

for i in range(len(name_list_off)):
    name = name_list_off[i]
    row_num= int(name[0])
    col_name = name[1]
    
    for j in range(len(row_cat)):
        df.loc[(row_num, row_cat[j]), col_name] = all_data_off[j][i]
        #all_data[j] is the list of a specific category (same with row_cat)
        #all_data[j][i] is the value in the list that corresponds to the name

        all_subdf_id = [(row,cat) for cat in row_cat for row in row_id]

all_index_off = [[],[],[],[]]

#Deal with index list from source dataframe
for idx in all_index_list_id:
    for i in chosen_index_list:
        all_index_off[idx].append(all_index_list[idx][i])
    
#Fill the empty spaces with NaN
df = df.replace(r'^\s*$', np.nan, regex=True)   

#Create id to create sub dataframes of each category
eff_subdf_id = all_subdf_id[:6]
ff_subdf_id = all_subdf_id[6:12]
voc_subdf_id = all_subdf_id[12:18]
jsc_subdf_id = all_subdf_id[18:24]

all_subdf_id = [eff_subdf_id, ff_subdf_id, voc_subdf_id, jsc_subdf_id]

#Create sub dataframes 
eff_subdf = df.loc[eff_subdf_id]
ff_subdf = df.loc[ff_subdf_id]
voc_subdf = df.loc[voc_subdf_id]
jsc_subdf = df.loc[jsc_subdf_id]

all_subdfs = [eff_subdf, ff_subdf, voc_subdf, jsc_subdf]

#Find best and worst value for each sub dataframe
best_list = []
worst_list = []
avg_list = []

for subdf in all_subdfs[:-1]:
    best_list.append(subdf.max().max())
    worst_list.append(subdf.min().min())
    avg_list.append(round(subdf.mean().mean(),3))
    
#Jsc values are negative: min - best value, max - worst value
best_list.append(jsc_subdf.min().min())
worst_list.append(jsc_subdf.max().max())
avg_list.append(round(subdf.mean().mean(),3))
    
#Find the index of the special values in the dataframe
best_df_id_list = []
for i in range(len(best_list)):
    best_id = find_index(best_list[i], df, row_cat[i])
    best_df_id_list.append(best_id)

worst_df_id_list = []
for i in range(len(worst_list)):
    worst_id = find_index(worst_list[i], df, row_cat[i])
    worst_df_id_list.append(worst_id)

#Find best cells names.
#Example of best_df_id_list = [[(1, 'Eff'), 'D'], [(1, 'Ff'), 'C'],\
#[(1, 'Voc'), 'D'], [(1, 'Jsc'), 'C']]

best_cells_names = [str(id[0][0]) + id[1] for id in best_df_id_list]
worst_cells_names = [str(id[0][0]) + id[1] for id in worst_df_id_list]

best_cell_df_id = best_cells_names[chosen_cat]

best_xl_id_list = []

for df_id in best_df_id_list:
    xl_id = xl_coordinate(df_id, df)
    best_xl_id_list.append(xl_id)
    
worst_xl_id_list = []
for df_id in worst_df_id_list:
    xl_id = xl_coordinate(df_id, df)
    worst_xl_id_list.append(xl_id)

#GRAPH THE CHAMPION CELL ------------------------------------------------------
"""
Find the champion cell name and position in raw dataframe
Extract data
And create graph 
"""
best_cat = best_list[chosen_cat]

#Find the best_cat in the original official list
best_cat_row_id = all_data_off[chosen_cat].index(best_cat)
best_cell = best_cells_names[chosen_cat]

#Match the data and the name to find voltage
while name_list_off[best_cat_row_id][:2] != best_cell:
    best_cat_row_id = all_data_off[chosen_cat].index(best_cat, \
                                                     best_cat_row_id + 1)

check_id = best_cat_row_id + 1

#for cases when cells with the same name and same data point, take the data of
#the most recent cell

while all_data_off[chosen_cat][check_id] == best_cat and \
      name_list_off[check_id][:2] == best_cell:
    best_cat_row_id = check_id
    check_id += 1

#best_cat_row_raw_df is the row of the best piece of information chosen
best_cat_row_raw_df = all_index_off[chosen_cat][best_cat_row_id]

#best_forward is a boolean to indicate if the data chosen is forward
best_forward = chosen_index_list[best_cat_row_id] % 2 == 0

#Find the data to plot about the champion cell

#Find the row 
#Find the voltage row associated with the best piece of information of the 
#chosen data

v_row = row_num_diff(best_cat_row_raw_df, chosen_cat)

if best_forward: #if the data is forward data
    
    #get the row, make it into a list, access the datetime object
    #and turn that into a string
    
    date = str(data.iloc[[v_row+11]].values.tolist()[0][0])[:10]
  
    v_forward_data = data.iloc[[v_row]].values.tolist()[0][1:]
    i_forward_data = data.iloc[[v_row+1]].values.tolist()[0][1:]
    
    v_reverse_data = data.iloc[[v_row+13]].values.tolist()[0][1:]
    i_reverse_data = data.iloc[[v_row+14]].values.tolist()[0][1:]

elif not best_forward: #if the data is reverse data
    
    date = str(data.iloc[[v_row-2]].values.tolist()[0][0])[:10]
    
    v_reverse_data = data.iloc[[v_row]].values.tolist()[0][1:]
    i_reverse_data = data.iloc[[v_row+1]].values.tolist()[0][1:] 
    
    v_forward_data = data.iloc[[v_row-13]].values.tolist()[0][1:]
    i_forward_data = data.iloc[[v_row-12]].values.tolist()[0][1:]

#Find the best_cell_name:
best_cell_name = name_list_off[best_cat_row_id][:8]

#Create_graph: Voltage - x axis, Current - y axis   
title = "IV curve - " + best_cell_name + " measured " + date
red = "#ff0000"
blue = "#0000ff"

plt.scatter(v_forward_data, i_forward_data, s = 2, label = "Forward", \
            color = blue, marker =',')
plt.scatter(v_reverse_data, i_reverse_data, s = 2, label = "Reverse", \
            color = red)
plt.title(title, fontweight = 'bold')
plt.xlabel('Voltage (V)')
plt.ylabel('Current (A)')
plt.legend()

#Add the axes through the origin
plt.axvline(x=0, color = 'k', linewidth = 0.5, alpha = 0.8)
plt.axhline(y=0, color = 'k', linewidth = 0.5, alpha = 0.8)

fig_file_name = best_cell_name + " measured " + date + '.png'
plt.savefig(fig_file_name)
plt.close()

#CREATE LEGEND ----------------------------------------------------------------
#Get the xl coordinate of the best cell based on the chosen characteristic
best_cell_xl_id = best_xl_id_list[chosen_cat]

df_analysis = pd.DataFrame({'Categories': row_cat,'Best': best_list, \
                            'Best Cell': best_cells_names, 'Worst': worst_list,\
                            'Worst Cell': worst_cells_names, 'Avg': avg_list, \
                            'Unit': units})

eff_style = eff_subdf.style.applymap(color_eff)

df_style = df.style.applymap(all_color_functions[chosen_cat],\
                             subset=pd.IndexSlice[all_subdf_id[chosen_cat], \
                                                  cols])

#Create color legend
all_df_leg = create_legend_df()
df_legend = all_df_leg[0]
df_legend_style = all_df_leg[1]

#EXPORT THE FILE --------------------------------------------------------------
#Get export file name
output_filename = get_export_filename(xl_sourcefile)
writer = ExcelWriter(output_filename) 

df_style.to_excel(writer,'Sheet1', index = True, startrow = 0, \
                      startcol = 0)
df_analysis.to_excel(writer, 'Sheet1', index = False, startrow = 0, \
                     startcol = len(df.columns) + 3)

df_legend_style.to_excel(writer,'Sheet1', index = False, 
                         startrow = len(df_analysis.index) + 3, \
                         startcol = len(df.columns) + 3)
                        
writer.save()

#FORMAT WORKBOOK ---------------------------------------------------------------
wb = openpyxl.load_workbook(output_filename)
ws = wb.worksheets[0]

row_num_best_xl = int(best_cell_xl_id[1:])

for row in color_range(row_num_best_xl, chosen_cat):
    ws[best_cell_xl_id[0] + str(row)].fill= PatternFill \
    (start_color = "e3c7e9", end_color = "e3c7e9", \
     fill_type = "solid") 

#Add chart to the output excel file
img = openpyxl.drawing.image.Image(fig_file_name)
ws.add_image(img, 'S'+str(len(df_analysis.index) + 5))

#FORMAT: change border, column width, add empty column, text alignment
#Bold the best value of the characteristic chosen by user
ws[best_cell_xl_id].font = Font(bold=True)
    
#Insert columns
xl_cols = list(range(3,14,2))
for col in xl_cols:
    ws.insert_cols(col)

#Change column width
xl_cols_w = list('DFHJLN')         
for col in xl_cols_w:
    ws.column_dimensions[col].width = 15

#Change small columns width
xl_cols_small = list('CEGIKM')         
for col in xl_cols_small:
    ws.column_dimensions[col].width = 1.5

#Make everything center
for row_cells in ws.iter_rows():
    for cell in row_cells:
        cell.alignment = Alignment(horizontal='center', vertical='center')

#Delete the name in the Color column
for row in range(9, 15):
    ws['P' + str(row)] = ""

#Unmerge the cells
start_row_range = list(range(2,25,4))
end_row_range = list(range(5, 26, 4))

for i in range(len(start_row_range)):
    ws.unmerge_cells('A'+str(start_row_range[i])+':A' + str(end_row_range[i]))
    
#Insert empty rows
rows_add = list(range(6, 30, 5))
for row in rows_add:
    ws.insert_rows(row)
    
#Merge the unmerged back
start_row_range = list(range(2, 30, 5))
end_row_range = list(range(5, 31, 5))

for i in range(len(start_row_range)):
    ws.merge_cells('A'+str(start_row_range[i])+':A' + str(end_row_range[i]))

#Move cells up to make the legend table intact
ws.move_range("P12:Q15", rows=-1, cols=0)
ws.move_range("P17:Q17", rows=-2, cols=0)

#Change border based on whether it's reverse or forward data
red = "ff0000"
blue = "0000ff"

border_up_forward = Border(left=Side(style='medium', color = blue), 
                     right=Side(style='medium', color = blue), 
                     top=Side(style='medium', color = blue))

border_mid_forward = Border(left=Side(style='medium', color = blue), 
                     right=Side(style='medium', color = blue))

border_bot_forward = Border(left=Side(style='medium', color = blue), 
                     right=Side(style='medium', color = blue), 
                     bottom=Side(style='medium', color = blue))      

border_up_reverse = Border(left=Side(style='medium', color = red), 
                     right=Side(style='medium', color = red), 
                     top=Side(style='medium', color = red))

border_mid_reverse = Border(left=Side(style='medium', color = red), 
                     right=Side(style='medium', color = red))

border_bot_reverse = Border(left=Side(style='medium', color = red), 
                     right=Side(style='medium', color = red), 
                     bottom=Side(style='medium', color = red))

border_forward = [border_up_forward, border_mid_forward, border_mid_forward, \
                  border_bot_forward]
border_reverse = [border_up_reverse, border_mid_reverse, border_mid_reverse, \
                  border_bot_reverse]

#Helper function for bordering
def xl_range(cell_name):
    """
    Find xl_range of all information about one cell based on the name 
    of the cell. Return all cells in the range. 
    (After adding columns in the middle) - A (df) - D(xl)
    """
    start = 2
    end = 5
    row_num = int(cell_name[0])
    
    start_row = start*(row_num-1) + row_num*2 + row_num - 1
        
    end_row = start_row + 3
    
    col = chr(ord(cell_name[1]) + 3 + ord(cell_name[1]) - 65)
    
    return [col+ str(start_row) for start_row in \
            list(range(start_row, start_row + 4))]

#Border every cell based on whether it's forward or reverse data
for cell_name in name_list_off:
    cell_range = xl_range(cell_name)
    if "Reverse" not in cell_name:
        for i in range(len(cell_range)):
            ws[cell_range[i]].border = border_forward[i]
    else: #If it's forward data
        for i in range(len(cell_range)):
            ws[cell_range[i]].border = border_reverse[i]        

#Range of cells: a tuple of tuples

wb.save(output_filename)
 
ext_time = time.time() - start_time

print("Extraction completed. Extraction took " + str(round(ext_time, 2)) + \
" seconds.")
print("File name is: " + output_filename + ".")