import openpyxl
from openpyxl.styles import Font, Color, Alignment
from openpyxl.styles import colors
import pandas as pd
from pandas import ExcelWriter
from pandas import ExcelFile
from pandas import Index
import time
import sys

"""
A program to extract all critical information about a batch of solar cells
and print the information to an excel file.
"""

#RETRIEVE DATA FROM DATA FRAME--------------------------------------------------
"""
List of category: Eff (Efficiency), Ff (Fill Factor), Voc, Jsc (Isc/Device Area)  
"""

def find_index(df, value):
    """
    Return a string indicating the index of value [row_index, column_index]
    in the dataframe df.
    """
    index = ""
    for col in list(df.columns):
        for row in list(df.index):
            if df.at[row, col] == value:
                index += (str(row))
                index += col
                
    return index
    
def row_num_difference(category):
    """
    Find row number from name of a category. 
    """
    category = category.lower()
    
    if category == "eff":
        row_num_diff = 3
    elif category == "ff":
        row_num_diff = 4
    elif category == "voc":
        row_num_diff = 5
    elif category == "isc" or category == "jsc":
        row_num_diff = 6
    return row_num_diff

def color_eff(val):
    """
    Color for Efficiency.
    """    
    val = float(val)
    color = "white"
    if val < 1:
        color = "#ff6961" #Red
    elif val < 3:
        color = "#fdfd96" #Yellow
    elif val < 6:
        color = "#9aee9a" #Green
    elif val > 6:
        color = "#c2e3f2" #Blue
    else: 
        color = "white"
    return 'background-color: %s' % color  

def color_ff(val):
    """
    Color for Fill Factor.
    """    
    val = float(val)
    color = "white"
    if val < 0.25:
        color = "#ff6961" #Red
    elif val < 0.35:
        color = "#fdfd96" #Yellow
    elif val < 0.4:
        color = "#9aee9a" #Green
    elif val > 0.4:
        color = "#c2e3f2" #Blue
    else: 
        color = "white"
    return 'background-color: %s' % color 

def color_voc(val):
    """
    Color for Voc.
    """    
    val = float(val)
    color = "white"
    if val < 0.8:
        color = "#ff6961" #Red
    elif val < 0.9:
        color = "#fdfd96" #Yellow
    elif val < 1:
        color = "#9aee9a" #Green
    elif val > 1:
        color = "#c2e3f2" #Blue
    else: 
        color = "white"
    return 'background-color: %s' % color 

def color_jsc(value):
    if pd.isnull(value):
        color = "white"
    else:
        color = "#ffebef"
    return 'background-color: %s' % color 

def leg_colors(value):
    """
    Color for legend.
    """
    if value == "Red":
        color = "#ff6961"
    elif value == "Yellow":
        color = "#fdfd96"
    elif value == "Green":
        color = "#9aee9a"
    elif value == "Blue":
        color = "#c2e3f2"
    elif value == "White":
        color = "white"
    return 'background-color: %s' % color

def xl_coordinate(df_coor, category):
    
    """
    Change 1 dataframe coordinate into excel coordinate
    eg. 7A - B8 (eff)
    eg. 7A - L8 (ff)
    eg. 7A - B26 (voc)
    eg. 7A - K26 (jsc)
    
    """
    category.lower()
    xl_coor = ""
    
    if category == "eff":
        xl_coor += chr(ord(df_coor[1]) + 1)
        xl_coor += str(int(df_coor[0]) + 1)
    elif category == "ff":
        xl_coor += chr(ord(df_coor[1]) + 10)
        xl_coor += str(int(df_coor[0]) + 1)
    elif category == "voc":
        xl_coor += chr(ord(df_coor[1]) + 1)
        xl_coor += str(int(df_coor[0]) + 19)
    elif category == "jsc":
        xl_coor += chr(ord(df_coor[1]) + 10)
        xl_coor += str(int(df_coor[0]) + 19) 
    
    return xl_coor
    
def analysis(dataf):
    """
    Find max, min, mean of whole batch. Find average of each column.
    Add to dataframe
    """
    #Data is a 6x6 dataframe
    all_cols = list('ABCDEF')
    all_rows = list(range(1,7))  
    num_cols = len(dataf.columns)
    num_rows = len(dataf.index)
    num_empty_cells = dataf.isnull().sum().sum()
    num_filled_cells = num_cols * num_rows - num_empty_cells
    
    #Find average of all batch
    all_mean = dataf.sum().sum()/num_filled_cells 
    
    dataf_abs = dataf.abs()
    
    #Find max
    max_cat = -100
    for col in all_cols:
        max_col = dataf_abs[col].max()
        if max_col > max_cat:
            max_cat = max_col
    
    #Find min
    min_cat = 100
    for col in all_cols:
        min_col = dataf_abs[col].min()
        if min_col < min_cat:
            min_cat = min_col
    
    #Find index name of max, min in the dataframe
    max_index = find_index(dataf, max_cat)
    min_index = find_index(dataf, min_cat)
    
    #Find average for each column and add to dataframe
    for col in all_cols:
        mean_col = dataf[col].mean()
        dataf.at[num_cols + 1, col] = round(mean_col, 3)
    
    #Find average for each row and add to dataframe
    for row in all_rows:
        mean_row = float(dataf.iloc[[row-1]].mean(axis=1))
        #axis = 1 because iloc returns a series in column
        #col - 1 because dataframe starts to count from 0        
        dataf.at[row, chr(num_rows + 1 + 64)] = round(mean_row, 3)
        #+64 because 'A' is 65
    
    #Print average at next column (G7 in dataframe)
    next_col = chr(ord(all_cols[-1])+1)
    dataf.at[num_rows + 1, next_col] = round(float(all_mean), 3)
    
    #Find max of the last row, which contains data about the average 
    #performance of each column
    max_col_val = dataf.iloc[[num_rows]].max(axis = 1) 
    max_col_val_id = find_index(dataf.iloc[[num_rows]], float(max_col_val))[:2]
    
    #Find max of the 2nd to last column, which contains data about the average 
    #performance of each row
    max_row_val = float(dataf.iloc[:,num_cols].max())
    row_index = str(Index(dataf.iloc[:,num_cols]).get_loc(max_row_val)+1)
    max_row_val_id =  row_index + list(dataf.columns)[-1]
    
    return [dataf, max_index, min_index, max_col_val_id, max_row_val_id]

def analysis_jsc(dataf):
    """
    Find max, min, mean of whole batch. Find average of each column.
    Add to dataframe
    """
    #Data is a 6x6 dataframe
    all_cols = list('ABCDEF')
    all_rows = list(range(1,7))  
    num_cols = len(dataf.columns)
    num_rows = len(dataf.index)
    num_empty_cells = dataf.isnull().sum().sum()
    num_filled_cells = num_cols * num_rows - num_empty_cells
    
    #Find average of all batch
    all_mean = dataf.sum().sum()/num_filled_cells 
    
    dataf_abs = dataf.abs()
    
    #Find max of absolute value but then keeps the value negative to 
    #find position
    max_cat_abs = -100
    for col in all_cols:
        max_col = dataf_abs[col].max()
        if max_col > max_cat_abs:
            max_cat_abs = max_col
    
    max_cat = -max_cat_abs
    
    
    #Find min
    min_cat_abs = 100
    for col in all_cols:
        min_col = dataf_abs[col].min()
        if min_col < min_cat_abs:
            min_cat_abs = min_col
    
    min_cat = -min_cat_abs
    
    #Find index name of max, min in the dataframe
    max_index = find_index(dataf, max_cat)
    min_index = find_index(dataf, min_cat)
    
    #Find average for each column and add to dataframe
    for col in all_cols:
        mean_col = dataf[col].mean()
        dataf.at[num_cols + 1, col] = round(mean_col, 3)
    
    #Find average for each row and add to dataframe
    for row in all_rows:
        mean_row = float(dataf.iloc[[row-1]].mean(axis=1))
        #axis = 1 because iloc returns a series in column
        #col - 1 because dataframe starts to count from 0        
        dataf.at[row, chr(num_rows + 1 + 64)] = round(mean_row, 3)
        #+64 because 'A' is 65
    
    #Print average at next column (G7 in dataframe)
    next_col = chr(ord(all_cols[-1])+1)
    dataf.at[num_rows + 1, next_col] = round(float(all_mean), 3)
    
    #Find max of the last row, which contains data about the average 
    #performance of each column
    max_col_val = dataf.iloc[[num_rows]].max(axis = 1) 
    max_col_val_id = find_index(dataf.iloc[[num_rows]], float(max_col_val))[:2]
    
    #Find max of the 2nd to last column, which contains data about the average 
    #performance of each row
    max_row_val = float(dataf.iloc[:,num_cols].max())
    row_index = str(Index(dataf.iloc[:,num_cols]).get_loc(max_row_val)+1)
    max_row_val_id =  row_index + list(dataf.columns)[-1]
    
    return [dataf, max_index, min_index, max_col_val_id, max_row_val_id]

def create_df(xl_source, category):
    """
    Create a dataframe of data of a specific category in the xl_file (Eff, Fill
    Factor, and Voc ONLY).
    Apply colors
    """
    category = category.lower()
    
    #Open the excel file
    wb = openpyxl.load_workbook(xl_source)
    sheet = wb.worksheets[0]    
      
    df = pd.read_excel(xl_source)
    
    name_list = []
    cat_index_list = [] #indexing in the dataframe 
    cat_list = []
    
    row_num_fr_name = row_num_difference(category)
    
    for index, row in df.iterrows():
        l_row = list(row)
        if "Comment:" in l_row:
            if "Reverse" not in l_row[1] and "d" not in l_row[1]:
                #l_row[1] is the name of the cell
                name_list.append(l_row[1])
                cat_index = index + row_num_fr_name
                cat_index_list.append(cat_index)
                                     
    #Indexing in excel is different 
    xl_cat_list = [str(i+2) for i in cat_index_list] 
    
    for row_num in xl_cat_list:
        cat_list.append(sheet['B' + row_num].value)        
                    
    #Create a new dataframe
    all_cols = list('ABCDEF')
    all_rows = list(range(1,7))
    data = pd.DataFrame(index = all_rows, columns = all_cols)
    
    num_cols = len(data.columns)
    num_rows = len(data.index)
    
    #Put data in the dataframe
    for i in range(len(name_list)):
        cell_name = name_list[i][:2]
        data.at[int(cell_name[0]), cell_name[1]] = cat_list[i]
    
    if category == "eff":
        data_style = data.style.applymap(color_eff)
    elif category == "ff":
        data_style = data.style.applymap(color_ff)
    elif category == "voc":
        data_style = data.style.applymap(color_voc)
    
    data_info = analysis(data)
    
    return [data_info[0],data_style, data_info[1:]]

def create_df_jsc(xl_source):
    """
    Create a dataframe for our baby Isc, which needs to be turned into 
    Current Density 
    """
    
    #Open the excel file
    wb = openpyxl.load_workbook(xl_source)
    sheet = wb.worksheets[0]     
    
    #Create a dataframe of all data
    df = pd.read_excel(xl_source)
    
    name_list = []
    isc_index_list = [] #indexing in the dataframe 
    isc_list = []
    
    area_index_list = []
    area_list = [] #please put in cm^2
    jsc_list = []
    
    row_num_fr_name = row_num_difference("isc")
    
    for index, row in df.iterrows():
        l_row = list(row)
        if "Comment:" in l_row:
            if "Reverse" not in l_row[1] and "d" not in l_row[1]:
                #l_row[1] is the name of the cell
                name_list.append(l_row[1])
                isc_index = index + row_num_fr_name
                isc_index_list.append(isc_index)
                
                area_index = index + 9
                area_index_list.append(area_index)
                       
    #Indexing in excel is different 
    xl_isc_list = [str(i+2) for i in isc_index_list] 
    xl_area_list = [str(i+2) for i in area_index_list]
    
    for row_num in xl_isc_list:
        isc_list.append(sheet['B' + row_num].value*1000)
        #change unit into mA
    
    for row_num in xl_area_list:
        area_list.append(sheet['B' + row_num].value)
    
    for i in range(len(isc_list)):
        jsc = isc_list[i]/area_list[i]
        jsc_list.append(jsc)
        
    #Create a new dataframe
    all_cols = list('ABCDEF')
    all_rows = list(range(1,7))      
    data = pd.DataFrame(index = all_rows, columns = all_cols)
    
    num_cols = len(data.columns)
    num_rows = len(data.index)
    
    #Put data in the dataframe
    for i in range(len(name_list)):
        cell_name = name_list[i][:2]
        data.at[int(cell_name[0]), cell_name[1]] = round(jsc_list[i], 3)
    
    data_style = data.style.applymap(color_jsc)
    data_info = analysis_jsc(data)
    
    return [data_info[0], data_style, data_info[1:]]

#Create legend
def create_legend_df():
    """
    Create the legend for the data
    """
    label_colors = ['Red', 'Yellow', 'Green', 'Blue', 'White']
    eff_values = ['< 1', '< 3', '< 6', '> 6', 'empty']
    ff_values = ['< 0.25', '< 0.35', '< 0.4', '> 0.4', 'empty']
    voc_values = ['< 0.8', '< 0.9', '< 1', '> 1', 'empty']
    jsc_values = ['', '', '', '', '']
    
    legend = {'Color': label_colors, 'Eff(%)': eff_values, 'FF': ff_values, \
              'Voc (V)': voc_values, 'Jsc (mA/mm^2)': jsc_values}
    
    df_legend = pd.DataFrame(legend)
    df_legend_style = df_legend.style.applymap(leg_colors, subset = ['Color'])
    
    return [df_legend,df_legend_style]
    
#EXPORT DATA ------------------------------------------------------------------
def get_export_filename(xl_source):
    """
    Get the name of cell and the date measured. Assume the name from the first
    cell measured and the date is from the last cell measured in the merged
    exel file.
    """
    wb = openpyxl.load_workbook(xl_source)
    sheet = wb.worksheets[0]
    
    #Get The Time Measured
    last_row_num = sheet.max_row
    
    #Most recent date indicates the data retrieved
    most_recent_date_row = last_row_num - 11
    
    #Get the date 
    date_measured = sheet["A" + str(most_recent_date_row)].value
    
    #Convert the date into string and get only the date
    date_str = str(date_measured).split()[0]     
    
    #Get the name of batch
    batch_name = sheet['B2'].value[3:]
    
    out_name = batch_name + " measured " + date_str + ".xlsx"
    
    return out_name

def export_data():
    source_file = input("What is the name of the source file? ") + ".xlsx"
    output_filename = get_export_filename(source_file)
    
    #Take all information about the dataframes
    all_df_eff = create_df(source_file, "eff")
    all_df_ff = create_df(source_file, "ff")
    all_df_voc = create_df(source_file, "voc")
    all_df_jsc = create_df_jsc(source_file)
    all_df_leg = create_legend_df()     
    
    #Create the four dataframes + dataframe for the legend
    df_eff = all_df_eff[0]
    df_ff = all_df_ff[0]
    df_voc = all_df_voc[0]
    df_jsc = all_df_jsc[0]
    df_leg_style = all_df_leg[0]
    
    #Create style objects to print to excel file
    df_eff_style = all_df_eff[1]
    df_ff_style = all_df_ff[1]
    df_voc_style = all_df_voc[1]
    df_jsc_style = all_df_jsc[1]
    df_leg_style = all_df_leg[1]
    
    #Export the data to an excel file in the same folder
    writer = ExcelWriter(output_filename)    
    
    #Paste the dataframes in the output excel file
    df_eff_style.to_excel(writer,'Sheet1', index = True, startrow = 0, \
                          startcol = 0)
    
    #Using the eff dataframe as a marker
    start_c = len(df_eff.columns)
    start_r = len(df_eff.index)
    df_ff_style.to_excel(writer, 'Sheet1', \
                   index = True, startrow = 0, startcol = start_c + 2)
    df_voc_style.to_excel(writer, 'Sheet1', index = True, \
                          startrow = start_r + 11, startcol = 0)
    df_jsc_style.to_excel(writer, 'Sheet1', index = True, startrow = start_r \
                          + 11, startcol = start_c + 2)
    df_leg_style.to_excel(writer, 'Sheet1', index = False, startrow = 10, \
                    startcol = 5)  
    writer.save()

    
    all_df = [df_eff, df_ff, df_voc, df_jsc]
    df_names = ["eff", "ff", "voc", "jsc"]
    all_id_list = [all_df_eff[2], all_df_ff[2], all_df_voc[2], all_df_jsc[2]]
    #all_df_category is a list including 
    #[max_index, min_index, max_col_val_id, max_row_val_id]
    
    all_special_coor = []
    #All special coor is a list of lists of special coordinates that
    #needs to be formatted
    
    #Convert the dataframe coordinates to xl coordinates to be formatted
    for i in range(len(all_df)-1):
        id_list = all_id_list[i]
        #id_list = [max_id, min_id, max_col_id,max_row_id] - dataf coordinates
        id_list_xl = []
        #xl_coordinates of id_list
        
        for coor in id_list:
            coor_xl = xl_coordinate(coor, df_names[i])
            id_list_xl.append(coor_xl)
        
        all_special_coor.append(id_list_xl)
    
    #Convert the dataframe coordinates to xl coordinates to be formatted for jsc
    id_list = all_df_jsc[2]
    #id_list = [max_id, min_id, max_col_id, max_row_id] - dataframe coordinates
    id_list_xl = []
    #xl_coordinates of id_list
    
    for coor in id_list:
        coor_xl = xl_coordinate(coor, "jsc")
        id_list_xl.append(coor_xl)
    
    all_special_coor.append(id_list_xl)  
    
    #Find the cell at the last row and the last column in the dataframes 
    #assuming the dataframes are similar.
    num_r = len(df_eff.index) 
    num_c = len(df_eff.columns) 
    
    row_id = str(num_r)
    col_id = chr(num_c + 64)
    
    avg_id = row_id + col_id  
    
    all_avg_xl_id = []
    # all_avg_xl_id is a list of all cells containing info about the average 
    #values of the whole batch
    
    for i in range(4):
        avg_xl_id = xl_coordinate(avg_id, df_names[i])
        all_avg_xl_id.append(avg_xl_id)
        
    #Format workbook
    wb = openpyxl.load_workbook(output_filename)
    ws = wb.worksheets[0]
    
    #Color the special values
    for special_coor_list in all_special_coor: 
        ws[special_coor_list[0]].font = Font(color = colors.RED)
        ws[special_coor_list[1]].font = Font(color = colors.GREEN)
        ws[special_coor_list[2]].font = Font(color = colors.RED)
        ws[special_coor_list[3]].font = Font(color = colors.RED)
        
    #Delete the name in the Color column
    for row in range(12, 17):
        ws['F' + str(row)] = ""
            
    #Label max and min
    ws['L11'] = 'Best'
    ws['L11'].font = Font(color = colors.RED)
    
    ws['L12'] = 'Worst'
    ws['L12'].font = Font(color = colors.GREEN)
    
    #Label the tables
    ws['A1'] = 'EFF'
    ws['J1'] = 'FF'
    ws['A19'] = 'Voc'
    ws['J19']= 'Jsc'
    
    #Format the names
    table_labels_pos= ['A1', 'J1', 'A19', 'J19']
    for pos in table_labels_pos:
        ws[pos].font = Font(bold = True)
        ws[pos].alignment = Alignment(horizontal='center')
    
    #Format the average cells
    for each_avg_id in all_avg_xl_id:
        ws[each_avg_id].font = Font(bold = True)
        
    wb.save(output_filename)   
    
    return output_filename
    
#PROMPT USER INPUT ------------------------------------------------------------
ques = "Have you converted the txt file into an excel file and saved it? [Y/N] "
ans = input(ques)

if ans != "Y" and ans != "y":
    print("Convert it and hit play button again.")
    sys.exit("Convert it and hit play button again.")

start_time = time.time()
output_filename = export_data()

ext_time = time.time() - start_time

print("Extraction completed. Extraction took " + str(round(ext_time, 2)) + \
" seconds.")

print("File name is: " + output_filename + ".")


